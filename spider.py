# -*- coding: utf-8 -*-
"""
Created on Tue Dec  25 9:30:01 2019

@author: 569273496
"""
from openpyxl import load_workbook
import datetime
import requests
import pandas as pd
import time




url = "http://quotes.money.163.com/service/chddata.html"
headers = {
            'User-Agent': "PostmanRuntime/7.20.1",
            'Accept': "*/*",
            'Cache-Control': "no-cache",
            'Postman-Token': "b4e16857-cb6f-4aea-b7f1-babb841fc269,96273727-a9c5-41d8-afd4-a9e013558759",
            'Host': "quotes.money.163.com",
            'Accept-Encoding': "gzip, deflate",
            'Connection': "keep-alive",
            'cache-control': "no-cache"
            }

def download(id,sheet):
    wb = load_workbook('code_dir\sh_code.xlsx')
    ws = wb[sheet]
    num=1
    while 1:
        
        if ws.cell(row=num, column=1).value:
            code=ws.cell(row=num, column=1).value
            date=ws.cell(row=num, column=3).value
            start=datetime.datetime.strptime(date,"%Y-%m-%d")
            end = start + datetime.timedelta(days=90)
            start=start.strftime("%Y%m%d")
            end=end.strftime("%Y%m%d")






            querystring = {"code":str(id)+str(code),"start":str(start),"end":str(end),"fields":"TCLOSE;HIGH;LOW;TOPEN;LCLOSE;CHG;PCHG;TURNOVER;VOTURNOVER;VATURNOVER;TCAP;MCAP"}
            response = requests.request("GET", url, headers=headers, params=querystring)
            text=response.text
            print('数据下载完毕','开始数据处理')
            
            csv='origin/'+str(code)+'.csv'
            # txt='origin/test.txt'
            new_csv='result/'+str(code)+'.csv'
            with open(csv, 'w',encoding='utf-8',newline='') as f:
                f.write(response.text)


            data =pd.read_csv(csv) 

            
            keep_col = [data.columns[0],data.columns[1],data.columns[2],data.columns[3],data.columns[4],data.columns[5],data.columns[6],data.columns[7],data.columns[8],data.columns[9],data.columns[10],data.columns[11]]
            # print(f[keep_col])

            new_f = data[keep_col]
            
            new_f.to_csv(new_csv, sep=',', header=True, index=False,encoding = 'utf_8_sig')
            sorf_csv=pd.read_csv(new_csv,  parse_dates=True).sort_index(ascending=False)  


            #去除停牌记录
            print('去除停牌数据')
            df2=sorf_csv.copy()
            for h in range(len(df2[data.columns[4]])) :

                if (df2[data.columns[6]])[h]==0.0:
                    df2.drop(index=h,inplace=True)
                    
            # print(df2)


            print(str(code)+"数据处理完毕！开始下一个代码数据下载与处理")
            print('——————————————————————————————————————————————')

            
            df2.to_csv(new_csv, sep=',', header=True, index=False,encoding = 'utf_8_sig')
                # start=s.strip('-')
            num=num+1
            
            time.sleep(3)

        else:
            exit()


download('0','sheet1')
download('1','sheet2')
